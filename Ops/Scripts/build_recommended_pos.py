"""Build dated, qty-filled raw goods PO Google Sheets from supplier templates.

Reads the "DRAFT SL Recommended.xlsx" draft (column V = Recommended PO qty),
copies each supplier template into a new dated folder in My Drive root,
fills C9 (P.O. Recommended Place = today) and the yellow Qty column,
then renames each file to the evaluated P.O. Reference.

Skips:
- KIT-* SKUs entirely (Palin gets a kit-component breakdown placed by hand later)
- Suppliers Brooke (Drill), Bill (Drill Bits), Zoe (Deluxe Brush), Alice (Liners),
  Nancy (Wipes), Sunny (Pro Beauty Box) — not yet in cycle

Special supplier rules:
- Sally: rows for EMPTY-GLASS-BOT-LID / -BRUSH / -INNER are auto-totalled
  to =SUM of the bottle qty column (one of each per bottle).

Usage:
  python build_recommended_pos.py                # all 11 supplier POs
  python build_recommended_pos.py Sally          # one supplier (case-insensitive)
  python build_recommended_pos.py --date 2026-06-02   # override date

Requires gcloud creds: `gcloud auth login --enable-gdrive-access` against
remy@glamrdip.com so `gcloud auth print-access-token` returns a Drive+Sheets-capable token.
"""
import subprocess, warnings, sys, json, argparse
from datetime import date, datetime
warnings.filterwarnings('ignore')

from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
import openpyxl


# --- Configuration ---
DRAFT_PATH = '/Users/remy-m4/Downloads/DRAFT SL Recommended.xlsx'
DRAFT_SHEET = 'Sheet1'
DRAFT_SKU_COL = 1          # A
DRAFT_SUPPLIER_COL = 3     # C
DRAFT_QTY_COL = 22         # V — Recommended PO qty
GCLOUD_ACCOUNT = 'remy@glamrdip.com'

# Draft -> Template SKU remap (handles naming-convention drift)
SKU_REMAP = {
    'ACC-TIP-ALM': 'LOOSE-TIP-ALM',
    'ACC-TIP-BAL': 'LOOSE-TIP-BAL',
    'ACC-TIP-COF': 'LOOSE-TIP-COF',
    'ACC-TIP-SQU': 'LOOSE-TIP-SQU',
    'ACC-TIP-STI': 'LOOSE-TIP-STI',
    'AU-EMPTY-JAR': 'AU-ACC-JAR',
    'UK/EU-EMPTY-JAR': 'UK/EU-ACC-JAR',
    'CA-EMPTY-JAR': 'CA-ACC-JAR',
    'UK/EU-EMPTY-REM-BOT-120': 'UK/EU-ACC-RE1-BOT',
    'AU-EMPTY-REM-BOT-120': 'AU-ACC-RE1-BOT',
    'CA-EMPTY-REM-BOT-120': 'CA-ACC-RE1-BOT',
    'UK/EU-EMPTY-REM-BOT-500': 'UK/EU-ACC-RE5-BOT',
    'AU-EMPTY-REM-BOT-500': 'AU-ACC-RE5-BOT',
    'CA-EMPTY-REM-BOT-500': 'CA-ACC-RE5-BOT',
    'ACC-NAI-MAT-002': 'ACCi-NAI-MAT-BOX',
}

# Per supplier: (name, template_id, qty_col_letter, category, company, sum_total_rows)
# sum_total_rows: list of (sku, dst_row, sum_start_row, sum_end_row) — the cell at dst_row
# becomes =SUM(qty_col{sum_start_row}:qty_col{sum_end_row}). One row per accessory that totals
# its parent SKUs (lids per jar, lid+inner per bottle, etc).
ALL_TEMPLATES = [
    ('Sally',    '1gUw76VOWly-O4tBTPMfauo7bUqnM7guiEBNcKUX6esY', 'D', 'Glass Bottles',   'Isay Nail', [
        ('EMPTY-GLASS-BOT-LID',   41, 14, 40),
        ('EMPTY-GLASS-BOT-BRUSH', 42, 14, 40),
        ('EMPTY-GLASS-BOT-INNER', 43, 14, 40),
    ]),
    ('Palin',    '18OFEgoljQAUb-T63czH8jUuOcvA9ChW4TtKmhDCXmpA', 'E', 'Packaging',       'Elephant Color Printing', []),
    ('Elise',    '1aXT7c8pql5rYWBrtQ7blIqTIppauVK-J18XEnXJMljc', 'E', 'Pro Files',       'Oslong', []),
    ('Heather',  '1fEhDASkiQ1ZNNZtHSRUppnmgLN6auap7QkuuOtJSLeU', 'E', 'Bubble Mailers',  'Suzhou Star New Material', []),
    ('Kay',      '1g3P2scXyYqa76BRMgg1xh7yqP15dF780rEOBmG5UCdo', 'E', 'Cuticle Presser', 'Yiwu Pinpin E-commerce Firm', []),
    ('Linda',    '16wQfpIRIM_M7weu8s2cYHZva0BMPF-XPlk3gKjAAqdU', 'F', 'Nail Tips',       'Yiwu Felice Cosmetic', []),
    ('Mark',     '1opvxBbQ_-zXbA2l3cmduIGYQwbzeWoxjGJpfpcoMQTQ', 'D', 'Jars',            'Yiwu Pinpin E-commerce Firm', [
        ('ACC-JAR-LIDS',  17, 14, 16),   # lids = SUM of all 3 jar regions
        ('ACC-JAR-SEALS', 18, 14, 16),
    ]),
    ('Michelle', '1RmLCHGPeNXOjixR4ESmuIQnrmnV68CX0aheJAtsMt4s', 'E', 'Mani Mat',        'Huizhou Yongli Industrial Co', []),
    ('Andy',     '1lFT8WNbkEuHhUNgK4O3xMiQv-jh28iVkiJi6uDfb8LA', 'E', 'Satchel',         'Shenzhen Hongxiang Packaging', []),
    ('Selena',   '1GDwWhtWHdLImOvgobQVJAy8yMhaiJ0D4HRs4uSBpCj4', 'D', 'Remove Bottles',  'Butuo', [
        ('ACC-RE1-LID', 17, 14, 16),     # 120ml lid   = SUM 120ml bottles (R14-R16)
        ('ACC-RE1-INN', 18, 14, 16),     # 120ml inner
        ('ACC-RE5-LID', 22, 19, 21),     # 500ml lid   = SUM 500ml bottles (R19-R21)
        ('ACC-RE5-INN', 23, 19, 21),     # 500ml inner
    ]),
    ('Bowls',    '13hU3j0L-HGl98JpZ-9aFZKWwbyMqU2UEvpnaozXZ7aw', 'E', 'Remove Bowl',     'Huzhou Jianaier International Trade', []),
    ('Zoe',      '1b5Zs719MGNx8SoDOmFrXtpc6cP_2rKxcc5yXusEQAsA', 'E', 'Deluxe Brush',    'Jiangxi Meichi Cosmetics', []),
]


def get_creds():
    token = subprocess.check_output(
        ['gcloud', 'auth', 'print-access-token', f'--account={GCLOUD_ACCOUNT}']
    ).decode().strip()
    return Credentials(token=token)


def load_draft():
    """Return {sku: (supplier, qty)} for rows with col V > 0."""
    wb = openpyxl.load_workbook(DRAFT_PATH, data_only=True)
    ws = wb[DRAFT_SHEET]
    out = {}
    for r in range(3, ws.max_row + 1):
        sku = ws.cell(r, DRAFT_SKU_COL).value
        supplier = ws.cell(r, DRAFT_SUPPLIER_COL).value
        v = ws.cell(r, DRAFT_QTY_COL).value
        if sku and isinstance(v, (int, float)) and v > 0:
            out[sku.strip()] = (supplier, int(v))
    return out


def get_template_sku_rows(sheets, sid):
    """Return {sku: row_number} reading column B starting at row 14."""
    rng = "'Recommended Purchase Order'!B14:B100"
    resp = sheets.spreadsheets().values().get(spreadsheetId=sid, range=rng).execute()
    rows = resp.get('values', [])
    out = {}
    for i, row in enumerate(rows):
        if row and row[0]:
            sku = row[0].strip()
            out[sku] = 14 + i
            # Order-swap fallback for UK/EU vs EU/UK
            if sku.startswith('UK/EU-'):
                out['EU/UK-' + sku[6:]] = 14 + i
            elif sku.startswith('EU/UK-'):
                out['UK/EU-' + sku[6:]] = 14 + i
    return out


def map_draft_to_template_row(draft_sku, tmpl_sku_rows):
    if draft_sku.startswith('KIT-'):
        return None, 'SKIPPED_KIT'
    if draft_sku in tmpl_sku_rows:
        return tmpl_sku_rows[draft_sku], 'EXACT'
    if draft_sku in SKU_REMAP and SKU_REMAP[draft_sku] in tmpl_sku_rows:
        return tmpl_sku_rows[SKU_REMAP[draft_sku]], f'REMAP({SKU_REMAP[draft_sku]})'
    return None, 'NOT_FOUND'


def ensure_folder(drive, folder_name):
    q = (f"name = '{folder_name}' and mimeType = 'application/vnd.google-apps.folder' "
         f"and trashed = false and 'root' in parents")
    existing = drive.files().list(q=q, fields='files(id,name,webViewLink)').execute()
    if existing.get('files'):
        f = existing['files'][0]
        print(f"Folder exists: {f['name']} ({f['id']})")
        return f['id']
    folder = drive.files().create(
        body={'name': folder_name, 'mimeType': 'application/vnd.google-apps.folder', 'parents': ['root']},
        fields='id,name,webViewLink').execute()
    print(f"Created folder: {folder['name']} -> {folder['webViewLink']}")
    return folder['id']


def build_po(drive, sheets, today, folder_id, cfg, draft):
    name, tmpl_id, qty_col, category, company, sum_rows = cfg
    print(f"\n--- {name} ---")
    tmpl_sku_rows = get_template_sku_rows(sheets, tmpl_id)

    fills = []
    for d_sku, (_supplier, qty) in draft.items():
        row, mt = map_draft_to_template_row(d_sku, tmpl_sku_rows)
        if row:
            fills.append((d_sku, row, qty, mt))

    if not fills:
        print("  No fills — skipping")
        return None

    date_str = today.strftime("%d-%m-%Y")
    tmp_name = f"{date_str} {name} (building)"
    copy_resp = drive.files().copy(
        fileId=tmpl_id,
        body={'name': tmp_name, 'parents': [folder_id]},
        fields='id,name,webViewLink',
        supportsAllDrives=True).execute()
    new_id = copy_resp['id']
    print(f"  Copied -> {new_id}")

    meta = sheets.spreadsheets().get(spreadsheetId=new_id).execute()
    sheet_id = meta['sheets'][0]['properties']['sheetId']
    qty_col_idx = ord(qty_col) - ord('A')
    date_serial = (today - date(1899, 12, 30)).days

    requests = [{
        'updateCells': {
            'range': {'sheetId': sheet_id, 'startRowIndex': 8, 'endRowIndex': 9,
                      'startColumnIndex': 2, 'endColumnIndex': 3},
            'rows': [{'values': [{'userEnteredValue': {'numberValue': date_serial}}]}],
            'fields': 'userEnteredValue'
        }
    }]
    for d_sku, row, qty, mt in fills:
        requests.append({
            'updateCells': {
                'range': {'sheetId': sheet_id, 'startRowIndex': row-1, 'endRowIndex': row,
                          'startColumnIndex': qty_col_idx, 'endColumnIndex': qty_col_idx+1},
                'rows': [{'values': [{'userEnteredValue': {'numberValue': qty}}]}],
                'fields': 'userEnteredValue'
            }
        })
        print(f"  {d_sku} -> row {row} qty {qty:,} [{mt}]")

    # Sum-total rows (lids/seals/inners that total their parent SKUs)
    for sum_sku, dst_row, start, end in sum_rows:
        formula = f'=SUM({qty_col}{start}:{qty_col}{end})'
        requests.append({
            'updateCells': {
                'range': {'sheetId': sheet_id, 'startRowIndex': dst_row-1, 'endRowIndex': dst_row,
                          'startColumnIndex': qty_col_idx, 'endColumnIndex': qty_col_idx+1},
                'rows': [{'values': [{'userEnteredValue': {'formulaValue': formula}}]}],
                'fields': 'userEnteredValue'
            }
        })
        print(f"  {sum_sku} -> row {dst_row} formula {formula}")

    sheets.spreadsheets().batchUpdate(spreadsheetId=new_id, body={'requests': requests}).execute()

    final_name = f"{date_str} | Raw Goods PO | {category} | {company}"
    drive.files().update(fileId=new_id, body={'name': final_name}, supportsAllDrives=True).execute()
    print(f"  Renamed -> {final_name}")

    return {
        'supplier': name,
        'file_id': new_id,
        'filename': final_name,
        'link': copy_resp['webViewLink'],
        'fills': fills,
    }


def main():
    p = argparse.ArgumentParser()
    p.add_argument('supplier', nargs='?', help='Run only one supplier (case-insensitive).')
    p.add_argument('--date', default=None, help='Override date YYYY-MM-DD (default today).')
    args = p.parse_args()

    today = date.fromisoformat(args.date) if args.date else date.today()
    date_str = today.strftime("%d-%m-%Y")
    print(f"Building POs for date {date_str}")

    creds = get_creds()
    drive = build('drive', 'v3', credentials=creds, cache_discovery=False)
    sheets = build('sheets', 'v4', credentials=creds, cache_discovery=False)

    draft = load_draft()
    print(f"Loaded {len(draft)} draft rows with V>0")

    folder_id = ensure_folder(drive, f"{date_str} Recommended POs")

    templates = ALL_TEMPLATES
    if args.supplier:
        templates = [t for t in ALL_TEMPLATES if t[0].lower() == args.supplier.lower()]
        if not templates:
            print(f"Unknown supplier '{args.supplier}'. Available: {[t[0] for t in ALL_TEMPLATES]}")
            return

    results = []
    for cfg in templates:
        r = build_po(drive, sheets, today, folder_id, cfg, draft)
        if r:
            results.append(r)

    print("\n========================================")
    print(f"Folder: https://drive.google.com/drive/folders/{folder_id}")
    for r in results:
        print(f"  - {r['filename']}")
        print(f"    {r['link']}")


if __name__ == '__main__':
    main()
