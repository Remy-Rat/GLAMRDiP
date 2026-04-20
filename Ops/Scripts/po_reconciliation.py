"""
PO Reconciliation — matches raw-good PO quantities against portal transfer history.

Inputs (positional args, or defaults below):
  1. PO Lots CSV export     (from portal: aggregate PO-level ordered/remaining)
  2. Transfer History XLSX  (from portal: SKU-level transfer ledger)

Output: multi-sheet XLSX at the path passed as the 3rd arg, or next to the inputs.

Sheets produced:
  - PO Summary          one row per PO: Ordered / Transferred / Implied Remaining / Portal Remaining / Variance
  - PO x SKU            per-SKU deductions against each PO (approved transfers only)
  - Allocation Map      every approved transfer, classified by FlowType (DIRECT/BUFFER_LEG1/BUFFER_LEG2)
  - CN Filing Summary   per CN Filing PO: SKUs + units + contributing source POs
  - Buffer Audit        reconciles BUFFER_LEG2 (CN-only) transfers against prior LEG1 (PO-only) legs per supplier+SKU
  - Transfer Raw        the source transfer rows, approved only, with FlowType tag

Flow types (per Greg):
  DIRECT_TO_CN       PO + CN Filing both set on transfer → deducts PO directly at shipment to Sally
  VIA_BUFFER_LEG1    PO set, CN blank → source supplier sends to intermediate, PO deducted here
  VIA_BUFFER_LEG2    PO blank, CN set → intermediate forwards to Sally, PO already deducted upstream
"""

import csv
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def classify_flow(po: str, cn: str, to: str = "") -> str:
    po, cn, to = (po or "").strip(), (cn or "").strip(), (to or "").strip()
    # Opening-balance seed entries use CN = CN-OPENING
    if cn.upper() == "CN-OPENING":
        return "OPENING_SEED"
    is_final = (to == "Sally")
    if po and cn:
        return "DIRECT_TO_CN" if is_final else "BUFFER_FEED_TAGGED"
    if po and not cn:
        return "BUFFER_FEED"  # source -> intermediate, deducts originator PO
    if cn and not po:
        return "BUFFER_FORWARD"  # intermediate -> Sally, CN assignment only
    return "UNCLASSIFIED"  # both blank — data anomaly


def parse_region(cn: str) -> str:
    if not cn:
        return ""
    up = cn.upper()
    for code in ("AUS", "UK", "CA", "NORDIC", "EU"):
        if up.startswith(code):
            return code
    return ""


def load_po_lots(path: Path):
    out = {}
    with open(path, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            po = r["PO Number"].strip()
            out[po] = {
                "supplier": r["Supplier"].strip(),
                "po_date": r["PO Date"].strip(),
                "invoice": r["Invoice"].strip(),
                "est_completion": r["Est.Completion"].strip(),
                "lines": int(r["Lines"] or 0),
                "ordered": int((r["Ordered"] or "0").replace(",", "")),
                "remaining_portal": int((r["Remaining"] or "0").replace(",", "")),
                "status": r["Status"].strip(),
            }
    return out


def load_transfers(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Transfer History"]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        req_id, status, frm, to, tdate, rdate, approver, adate, sku, po, cn, units = r
        if status != "APPROVED":
            continue
        rows.append({
            "req_id": req_id, "status": status, "from": frm, "to": to,
            "transfer_date": str(tdate) if tdate else "",
            "request_date": str(rdate) if rdate else "",
            "approver": approver,
            "approval_date": str(adate) if adate else "",
            "sku": sku, "po": (po or "").strip(), "cn": (cn or "").strip(),
            "units": int(units or 0),
            "flow": classify_flow(po, cn, to),
        })
    return rows


HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
FLAG_OK_FILL = PatternFill("solid", fgColor="DCFCE7")
FLAG_WARN_FILL = PatternFill("solid", fgColor="FEE2E2")


def write_sheet(wb, title, headers, rows, flag_col=None):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left")
    for row in rows:
        ws.append(row)
        if flag_col is not None:
            v = row[flag_col - 1]
            if v == "OK":
                ws.cell(row=ws.max_row, column=flag_col).fill = FLAG_OK_FILL
            elif v and v != "":
                ws.cell(row=ws.max_row, column=flag_col).fill = FLAG_WARN_FILL
    # autosize cols (rough)
    for i, h in enumerate(headers, 1):
        width = max(len(str(h)), 10)
        for row in rows[:200]:
            v = row[i - 1] if i - 1 < len(row) else ""
            width = max(width, min(len(str(v)), 45))
        ws.column_dimensions[get_column_letter(i)].width = width + 2
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return ws


def build_po_summary(pos, transfers):
    # deducted by PO = sum of LEG1 + DIRECT transfers (PO is set)
    deducted = defaultdict(int)
    for t in transfers:
        if t["po"]:
            deducted[t["po"]] += t["units"]

    rows = []
    for po, meta in pos.items():
        ded = deducted.get(po, 0)
        implied_remaining = meta["ordered"] - ded
        portal_remaining = meta["remaining_portal"]
        variance = portal_remaining - implied_remaining
        flag = "OK" if variance == 0 else f"MISMATCH ({variance:+d})"
        rows.append([
            po, meta["supplier"], meta["po_date"], meta["invoice"],
            meta["est_completion"], meta["status"],
            meta["ordered"], ded, implied_remaining, portal_remaining,
            variance, flag,
        ])
    # sort: mismatches first, then by supplier then PO
    rows.sort(key=lambda r: (r[11] == "OK", r[1], r[0]))
    return rows


def build_po_sku(transfers):
    agg = defaultdict(lambda: {"units": 0, "n": 0, "first": "", "last": ""})
    for t in transfers:
        if not t["po"]:
            continue
        key = (t["po"], t["sku"])
        a = agg[key]
        a["units"] += t["units"]
        a["n"] += 1
        d = t["transfer_date"]
        if d:
            a["first"] = d if not a["first"] or d < a["first"] else a["first"]
            a["last"] = d if not a["last"] or d > a["last"] else a["last"]
    rows = []
    for (po, sku), a in agg.items():
        rows.append([po, sku, a["units"], a["n"], a["first"], a["last"]])
    rows.sort(key=lambda r: (r[0], r[1]))
    return rows


def build_allocation_map(transfers):
    rows = []
    for t in transfers:
        rows.append([
            t["transfer_date"], t["req_id"], t["flow"],
            t["from"], t["to"], t["sku"], t["units"],
            t["po"], t["cn"], parse_region(t["cn"]),
        ])
    rows.sort(key=lambda r: (r[7], r[0], r[5]))  # PO, date, SKU
    return rows


def build_cn_summary(transfers):
    # aggregate by CN + SKU; capture source POs
    agg = defaultdict(lambda: {"units": 0, "pos": set(), "suppliers": set(), "legs": 0})
    for t in transfers:
        if not t["cn"]:
            continue
        key = (t["cn"], t["sku"])
        a = agg[key]
        a["units"] += t["units"]
        a["legs"] += 1
        if t["po"]:
            a["pos"].add(t["po"])
        a["suppliers"].add(t["from"])
    rows = []
    for (cn, sku), a in agg.items():
        rows.append([
            cn, parse_region(cn), sku, a["units"], a["legs"],
            ", ".join(sorted(a["suppliers"])),
            ", ".join(sorted(a["pos"])) if a["pos"] else "(buffer only)",
        ])
    rows.sort(key=lambda r: (r[0], r[2]))
    return rows


def build_buffer_audit(transfers):
    """Per (intermediate supplier, SKU): verify BUFFER_FORWARD outflows are backed
    by prior inflows. Counts all inbound flows (BUFFER_FEED, BUFFER_FEED_TAGGED,
    DIRECT_TO_CN, plus OPENING_SEED / UNCLASSIFIED) at the intermediate."""
    feed_in = defaultdict(int)       # BUFFER_FEED + BUFFER_FEED_TAGGED: standard inflows
    direct_in = defaultdict(int)     # DIRECT_TO_CN received (Sally mainly)
    other_in = defaultdict(int)      # OPENING_SEED / UNCLASSIFIED inflows
    fwd_out = defaultdict(int)       # BUFFER_FORWARD sent
    for t in transfers:
        k_in = (t["to"], t["sku"])
        k_out = (t["from"], t["sku"])
        if t["flow"] in ("BUFFER_FEED", "BUFFER_FEED_TAGGED"):
            feed_in[k_in] += t["units"]
        elif t["flow"] == "DIRECT_TO_CN":
            direct_in[k_in] += t["units"]
        elif t["flow"] in ("OPENING_SEED", "UNCLASSIFIED"):
            other_in[k_in] += t["units"]
        if t["flow"] == "BUFFER_FORWARD":
            fwd_out[k_out] += t["units"]

    rows = []
    keys = set(feed_in) | set(fwd_out)
    for k in list(direct_in) + list(other_in):
        if k in fwd_out:
            keys.add(k)

    for (supplier, sku) in sorted(keys):
        fi = feed_in.get((supplier, sku), 0)
        di = direct_in.get((supplier, sku), 0)
        oi = other_in.get((supplier, sku), 0)
        out = fwd_out.get((supplier, sku), 0)
        total_in = fi + di + oi
        holding = total_in - out
        if out == 0:
            flag = f"BUFFER HOLDING ({fi + oi} units at {supplier}, no forward yet)" if (fi or oi) else "OK"
        elif total_in == 0 and out > 0:
            flag = f"ORPHAN FORWARD (no recorded inflow for {supplier}/{sku})"
        elif out > total_in:
            flag = f"OVERSEND ({out - total_in} units beyond recorded inflow)"
        elif holding > 0:
            flag = f"BUFFER HOLDING ({holding} units at {supplier})"
        else:
            flag = "OK"
        rows.append([supplier, sku, fi, di, oi, out, holding, flag])
    rows.sort(key=lambda r: (r[7] == "OK", r[0], r[1]))
    return rows


def build_raw(transfers):
    rows = []
    for t in transfers:
        rows.append([
            t["transfer_date"], t["req_id"], t["flow"],
            t["from"], t["to"], t["sku"], t["units"],
            t["po"], t["cn"], parse_region(t["cn"]),
            t["request_date"], t["approval_date"], t["approver"],
        ])
    rows.sort(key=lambda r: r[0])
    return rows


def main():
    args = sys.argv[1:]
    po_lots_path = Path(args[0]) if len(args) >= 1 else Path("/Users/remy-m4/Downloads/PO_Lots_2026-04-17.csv")
    transfer_path = Path(args[1]) if len(args) >= 2 else Path("/Users/remy-m4/Downloads/Transfer_History_2026-04-16 (1).xlsx")
    out_path = Path(args[2]) if len(args) >= 3 else Path(
        f"/Users/remy-m4/Documents/GD/GLAMRDiP/Ops/Suppliers/PO_Reconciliation_{date.today().isoformat()}.xlsx"
    )

    print(f"Loading PO lots: {po_lots_path}")
    pos = load_po_lots(po_lots_path)
    print(f"  {len(pos)} POs loaded")

    print(f"Loading transfers: {transfer_path}")
    transfers = load_transfers(transfer_path)
    print(f"  {len(transfers)} approved transfer rows loaded")

    # check unknown POs in transfers
    unknown = sorted({t['po'] for t in transfers if t['po'] and t['po'] not in pos})
    if unknown:
        print(f"  WARNING: {len(unknown)} transfer POs not in PO Lots: {unknown}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet: PO Summary
    write_sheet(
        wb, "PO Summary",
        ["PO Number", "Supplier", "PO Date", "Invoice", "Est. Completion", "Status",
         "Ordered", "Transferred (calc)", "Implied Remaining", "Portal Remaining",
         "Variance (Portal - Implied)", "Flag"],
        build_po_summary(pos, transfers),
        flag_col=12,
    )

    # Sheet: PO x SKU
    write_sheet(
        wb, "PO x SKU",
        ["PO Number", "SKU", "Units Deducted", "# Transfers", "First Transfer", "Last Transfer"],
        build_po_sku(transfers),
    )

    # Sheet: Allocation Map
    write_sheet(
        wb, "Allocation Map",
        ["Transfer Date", "Request Id", "Flow Type", "From", "To", "SKU", "Units",
         "PO Number", "CN Filing", "Region"],
        build_allocation_map(transfers),
    )

    # Sheet: CN Filing Summary
    write_sheet(
        wb, "CN Filing Summary",
        ["CN Filing PO", "Region", "SKU", "Units", "# Legs", "Supplier(s)", "Source PO(s)"],
        build_cn_summary(transfers),
    )

    # Sheet: Buffer Audit
    write_sheet(
        wb, "Buffer Audit",
        ["Buffer Supplier", "SKU", "Feed In (PO-tagged)", "Direct In (PO+CN)",
         "Other In (Opening/Anomaly)", "Forward Out (CN-only)",
         "Implied Holding", "Flag"],
        build_buffer_audit(transfers),
        flag_col=8,
    )

    # Sheet: Transfer Raw
    write_sheet(
        wb, "Transfer Raw",
        ["Transfer Date", "Request Id", "Flow Type", "From", "To", "SKU", "Units",
         "PO Number", "CN Filing", "Region",
         "Request Date", "Approval Date", "Approver"],
        build_raw(transfers),
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"\nWrote: {out_path}")


if __name__ == "__main__":
    main()
